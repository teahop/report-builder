"""20-run History writer sweep on a frozen ledger. Writes JSONL + coding xlsx.

Does not re-extract. Does not retune. Does not accept the ledger.
Workbook matches docs/engineering/eval-templates/Open-coding workbook — TEMPLATE.xlsx.
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import subprocess
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

_WEEK1 = Path(__file__).resolve().parents[2]
if str(_WEEK1) not in sys.path:
    sys.path.insert(0, str(_WEEK1))

from dotenv import load_dotenv

load_dotenv(_WEEK1 / ".env")

from conflicts import detect_disagreements_from_ledger
from history_langfuse import flush_langfuse
from history_structure import structure_spec_hash
from history_writer import (
    DIAGNOSTIC_BANNER,
    TRACE_ALIGNMENT_STATUS,
    WriterSectionOutput,
    attach_plan_identity,
    plan_writer_calls,
    render_diagnostic_section,
    writer_prompt_hash,
)
from langfuse import observe
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from provider import BASTION_MODEL, DRAFT_TEMPERATURE, ModelProvider
from schemas import Ledger

_LEDGER = (
    _WEEK1
    / "evals"
    / "history"
    / "diagnostic_ladder"
    / "run-20260814T180644Z"
    / "ledger.json"
)
_TEMPLATE = (
    Path("/Users/thopk/ai-eng-bootcamp/docs/engineering/eval-templates")
    / "Open-coding workbook — TEMPLATE.xlsx"
)
_TRACES = _WEEK1 / "evals" / "traces"
_CODING = _WEEK1 / "evals" / "coding"
_N_RUNS = 20
_FIXTURE_ID = "fixture_001"
_QUOTE_RE = re.compile(r'"([^"]{2,})"')
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
_BAND_FILL = PatternFill("solid", fgColor="F7F7F7")
_HEADER_FONT = Font(name="Arial", size=10, bold=True, color="1F3864")
_BODY_FONT = Font(name="Arial", size=10)


def _sha_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _git_sha() -> str | None:
    try:
        return subprocess.check_output(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=_WEEK1,
            stderr=subprocess.DEVNULL,
            text=True,
        ).strip()
    except (OSError, subprocess.CalledProcessError):
        return None


def _load_ledger() -> Ledger:
    raw = json.loads(_LEDGER.read_text(encoding="utf-8"))
    return Ledger.model_validate(raw["ledger"] if "ledger" in raw else raw)


def _completed_indexes(path: Path) -> set[int]:
    if not path.is_file():
        return set()
    done: set[int] = set()
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        row = json.loads(line)
        if row.get("status") == "ok":
            done.add(int(row["run_index"]))
    return done


def _strip_banner(text: str) -> str:
    body = text
    if body.startswith(DIAGNOSTIC_BANNER):
        body = body[len(DIAGNOSTIC_BANNER) :].lstrip("\n")
    return body.strip() + "\n"


def _langfuse_ids() -> tuple[str | None, str | None]:
    try:
        from langfuse import get_client

        lf = get_client()
        trace_id = lf.get_current_trace_id()
        url = lf.get_trace_url(trace_id=trace_id) if trace_id else None
        return trace_id, url
    except Exception:
        return None, None


def _update_trace(eval_run_id: str, run_index: int, metadata: dict[str, Any]) -> None:
    try:
        from langfuse import get_client

        get_client().update_current_trace(
            name="history_writer_sweep_draft",
            session_id=eval_run_id,
            tags=["eval", "history", "positive_writer", "bastion", _FIXTURE_ID],
            metadata=metadata,
        )
    except Exception:
        pass


@observe(name="eval.history.writer_sweep.draft")
def _draft_one(
    provider: ModelProvider,
    *,
    model: str,
    plan: Any,
    requests: list[dict[str, Any]],
    eval_run_id: str,
    run_index: int,
    ledger_sha: str,
) -> dict[str, Any]:
    _update_trace(
        eval_run_id,
        run_index,
        {
            "fixture_id": _FIXTURE_ID,
            "run_index": run_index,
            "ledger_sha": ledger_sha,
            "package": "positive_history_writer",
            "confirm_synthetic": True,
        },
    )
    tokens_used = prompt_tokens = completion_tokens = 0
    assembled_parts: list[str] = []
    sections: list[dict[str, Any]] = []
    all_blocks: list[dict[str, Any]] = []
    start = time.perf_counter()
    for req in requests:
        result = provider.complete_structured(
            model=model,
            messages=req["messages"],
            schema=WriterSectionOutput,
            temperature=DRAFT_TEMPERATURE,
        )
        tokens_used += result.total_tokens
        prompt_tokens += result.prompt_tokens
        completion_tokens += result.completion_tokens
        output = result.data
        assert isinstance(output, WriterSectionOutput)
        section = next(s for s in plan.sections if s.section_key == req["section_key"])
        attached = attach_plan_identity(output, section, req["offered_evidence_ids"])
        readable = render_diagnostic_section(req["display_label"], attached)
        assembled_parts.append(readable)
        sections.append(
            {
                "section_key": req["section_key"],
                "display_label": req["display_label"],
                "blocks": attached,
                "prose": readable,
            }
        )
        for block in attached:
            all_blocks.append(
                {
                    "section_key": req["section_key"],
                    "section_label": req["display_label"],
                    "block_key": block["block_key"],
                    "block_label": block["display_label"],
                    "prose": block.get("prose") or "",
                    "fact_ids": list(block.get("offered_evidence_ids") or []),
                    "kind": block.get("kind"),
                }
            )
    latency_ms = int((time.perf_counter() - start) * 1000)
    assembled = "\n".join(assembled_parts).rstrip() + "\n"
    prose = _strip_banner(assembled) if assembled.startswith(DIAGNOSTIC_BANNER) else assembled
    # assembled_parts have no banner; keep them as the readable draft.
    quotes = _QUOTE_RE.findall(prose)
    trace_id, langfuse_url = _langfuse_ids()
    if not trace_id:
        trace_id = uuid4().hex
    return {
        "trace_id": trace_id,
        "langfuse_url": langfuse_url,
        "eval_run_id": eval_run_id,
        "fixture_id": _FIXTURE_ID,
        "run_index": run_index,
        "status": "ok",
        "package": "positive_history_writer",
        "provider": "bastion",
        "model": model,
        "temperature": DRAFT_TEMPERATURE,
        "writer_prompt_hash": writer_prompt_hash("Current Status & History"),
        "structure_spec_id": "provisional_tj_v1",
        "structure_spec_hash": structure_spec_hash("provisional_tj_v1"),
        "ledger_path": str(_LEDGER.relative_to(_WEEK1)),
        "ledger_sha": ledger_sha,
        "trace_alignment_status": TRACE_ALIGNMENT_STATUS,
        "skip_entailment": True,
        "section_populated": True,
        "prose": prose,
        "block_labels": " | ".join(b["block_label"] for b in all_blocks),
        "blocks": all_blocks,
        "sections": sections,
        "unanchored_quotes": quotes,
        "tokens_used": tokens_used,
        "tokens_by_stage": {"draft": tokens_used, "entailment": 0},
        "prompt_tokens": prompt_tokens,
        "completion_tokens": completion_tokens,
        "cost_usd": None,
        "latency_ms": latency_ms,
        "git_sha": _git_sha(),
    }


def _style_input_cell(cell: Any, *, wrap: bool = True) -> None:
    cell.font = _BODY_FONT
    cell.fill = _INPUT_FILL
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")


def _style_body_cell(cell: Any, *, band: bool, wrap: bool = True) -> None:
    cell.font = _BODY_FONT
    if band:
        cell.fill = _BAND_FILL
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")


def export_workbook(rows: list[dict[str, Any]], dest: Path) -> None:
    if not _TEMPLATE.is_file():
        raise FileNotFoundError(f"missing coding template: {_TEMPLATE}")
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(_TEMPLATE, dest)
    wb = load_workbook(dest)

    drafts = wb["Drafts"]
    for row in drafts.iter_rows(min_row=2, max_row=max(drafts.max_row, 23)):
        for cell in row:
            cell.value = None
            cell.hyperlink = None

    for i, record in enumerate(rows, start=1):
        r = i + 1
        band = i % 2 == 0
        drafts.row_dimensions[r].height = 210
        values = {
            "A": i,
            "B": record["trace_id"],
            "C": record.get("langfuse_url") or "",
            "D": record["fixture_id"],
            "E": record["run_index"],
            "F": record.get("block_labels") or "",
            "G": record["prose"],
            "H": len(record.get("unanchored_quotes") or []),
            "I": None,
            "J": None,
            "K": None,
        }
        for col, value in values.items():
            cell = drafts[f"{col}{r}"]
            cell.value = value
            if col in {"I", "J", "K"}:
                _style_input_cell(cell)
            else:
                _style_body_cell(cell, band=band, wrap=col in {"F", "G"})
        drafts[f"A{r}"].font = _BODY_FONT
        url = record.get("langfuse_url")
        if url:
            drafts[f"C{r}"].hyperlink = url
            drafts[f"C{r}"].font = Font(name="Arial", size=10, color="0563C1", underline="single")

    for cell in drafts[1]:
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    blocks = wb["Blocks"]
    for row in blocks.iter_rows(min_row=2, max_row=max(blocks.max_row, 10)):
        for cell in row:
            cell.value = None

    n = 0
    for record in rows:
        for block in record.get("blocks") or []:
            n += 1
            r = n + 1
            band = n % 2 == 0
            blocks.row_dimensions[r].height = 56
            fact_ids = block.get("fact_ids") or []
            values = {
                "A": n,
                "B": record["trace_id"],
                "C": record["fixture_id"],
                "D": record["run_index"],
                "E": block.get("block_label") or "",
                "F": block.get("prose") or "",
                "G": ", ".join(fact_ids),
                "H": None,
                "I": None,
                "J": None,
            }
            for col, value in values.items():
                cell = blocks[f"{col}{r}"]
                cell.value = value
                if col in {"H", "I", "J"}:
                    _style_input_cell(cell)
                else:
                    _style_body_cell(cell, band=band, wrap=col in {"E", "F"})
    for cell in blocks[1]:
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    tally = wb["Tally"]
    tally["C8"].number_format = FORMAT_PERCENTAGE_00
    for coord in ("C11", "C12", "C13", "C14", "C15"):
        tally[coord].number_format = FORMAT_PERCENTAGE_00

    wb.save(dest)


def main() -> int:
    if not _LEDGER.is_file():
        print(f"Missing Bastion ladder ledger: {_LEDGER}", file=sys.stderr)
        return 2
    if not _TEMPLATE.is_file():
        print(f"Missing coding template: {_TEMPLATE}", file=sys.stderr)
        return 2

    stamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    eval_run_id = f"sweep-{stamp}-{uuid4().hex[:6]}"
    _TRACES.mkdir(parents=True, exist_ok=True)
    _CODING.mkdir(parents=True, exist_ok=True)
    jsonl_path = _TRACES / f"{eval_run_id}.jsonl"
    xlsx_path = _CODING / f"{eval_run_id}.xlsx"

    ledger = _load_ledger()
    ledger_sha = _sha_file(_LEDGER)
    conflicts, variance, *_ = detect_disagreements_from_ledger(ledger)
    plan, requests = plan_writer_calls(ledger, conflicts=conflicts, variance=variance)
    provider = ModelProvider(backend="bastion")
    model = BASTION_MODEL

    print(
        f"20-run positive History writer sweep  provider=bastion model={model} "
        f"n={_N_RUNS} fixture={_FIXTURE_ID}"
    )
    print(f"eval_run_id={eval_run_id}")
    print(f"ledger={_LEDGER.relative_to(_WEEK1)} sha={ledger_sha[:12]}")
    print(f"writer_calls_per_run={len(requests)} total_writer_calls={len(requests) * _N_RUNS}")
    print("Does not re-extract. Ledger not accepted. Coding columns empty.")
    print()

    done = _completed_indexes(jsonl_path)
    records: list[dict[str, Any]] = []
    if jsonl_path.is_file():
        for line in jsonl_path.read_text(encoding="utf-8").splitlines():
            if line.strip():
                row = json.loads(line)
                if row.get("status") == "ok":
                    records.append(row)

    try:
        with jsonl_path.open("a", encoding="utf-8") as fh:
            for run_index in range(1, _N_RUNS + 1):
                if run_index in done:
                    print(f"  skip run {run_index} (already in jsonl)", flush=True)
                    continue
                print(f"  draft run {run_index}/{_N_RUNS} …", flush=True)
                record = _draft_one(
                    provider,
                    model=model,
                    plan=plan,
                    requests=requests,
                    eval_run_id=eval_run_id,
                    run_index=run_index,
                    ledger_sha=ledger_sha,
                )
                fh.write(json.dumps(record, ensure_ascii=False) + "\n")
                fh.flush()
                records.append(record)
                print(
                    f"    tokens={record['tokens_used']} latency_ms={record['latency_ms']} "
                    f"trace={record['trace_id'][:12]}",
                    flush=True,
                )
    finally:
        flush_langfuse()

    records.sort(key=lambda r: int(r["run_index"]))
    export_workbook(records, xlsx_path)
    print()
    print(f"jsonl={jsonl_path}")
    print(f"coding workbook={xlsx_path}")
    print(f"rows={len(records)}")
    print("HARD STOP — open-coding workbook is the review file; input columns empty.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
